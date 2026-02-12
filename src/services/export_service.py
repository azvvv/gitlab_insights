from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any  # 添加 Any 的导入
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database.connection import get_db_session
from database.models import GitlabRepositoryBranch, GitlabRepository, GitlabBranchRule
from dto.branch_dto import BranchExportData 
from sqlalchemy.orm import joinedload

class ExportService:
    def __init__(self):
        pass
    
    def _apply_branch_rules(self, branch: GitlabRepositoryBranch, db) -> Dict[str, Any]:
        """根据分支规则计算分支属性"""
        # 获取所有启用的分支规则，按优先级排序
        rules = db.query(GitlabBranchRule).filter(
            GitlabBranchRule.is_active == True
        ).order_by(GitlabBranchRule.priority.desc()).all()
        
        result = {
            'branch_type': branch.branch_type,  # 默认使用数据库中的值
            'is_deletable': False,
            'retention_deadline': None,
            'deletion_reason': None,
            'matched_rule_name': None
        }
        
        # 遍历规则，找到第一个匹配的规则
        for rule in rules:
            if self._branch_matches_pattern(branch.branch_name, rule.branch_pattern):
                result['branch_type'] = rule.branch_type
                result['is_deletable'] = rule.is_deletable and not branch.protected
                result['matched_rule_name'] = rule.rule_name
                
                # 计算保留截止时间
                if rule.retention_days and branch.last_commit_date:
                    result['retention_deadline'] = branch.last_commit_date + timedelta(days=rule.retention_days)
                    
                    # 检查是否已过期
                    if result['retention_deadline'] and datetime.now() > result['retention_deadline']:
                        result['deletion_reason'] = f"根据规则'{rule.rule_name}'，分支已超过{rule.retention_days}天保留期限"
                    elif result['is_deletable']:
                        result['deletion_reason'] = f"根据规则'{rule.rule_name}'，该类型分支可删除"
                elif result['is_deletable']:
                    result['deletion_reason'] = f"根据规则'{rule.rule_name}'，该类型分支可删除"
                
                # 受保护的分支不可删除
                if branch.protected:
                    result['is_deletable'] = False
                    result['deletion_reason'] = "分支受保护，不可删除"
                
                break  # 找到第一个匹配的规则后退出
        
        # 如果没有匹配的规则但分支受保护
        if branch.protected and not result['matched_rule_name']:
            result['deletion_reason'] = "分支受保护，不可删除"
        
        return result
    
    def _branch_matches_pattern(self, branch_name: str, pattern: str) -> bool:
        """检查分支名称是否匹配规则模式"""
        # 支持通配符匹配
        if pattern == '*':
            return True
        
        # 支持多个模式用|分隔
        patterns = pattern.split('|')
        for p in patterns:
            p = p.strip()
            # 将通配符模式转换为正则表达式
            regex_pattern = p.replace('*', '.*').replace('?', '.')
            if re.match(f'^{regex_pattern}$', branch_name, re.IGNORECASE):
                return True
        
        return False
    
    def export_branch_deletion_report_to_excel(self, repository_id: int = None) -> io.BytesIO:
        """导出分支删除报告为 Excel 文件"""
        try:
            print(f"Starting Excel export for repository_id: {repository_id}")
            
            # 获取并转换数据为 DTO
            export_data = self._get_branch_export_data(repository_id)
            print(f"Found {len(export_data)} branches converted to DTO")
            
            # 使用 DTO 创建 Excel，完全独立于数据库会话
            return self._create_excel_from_dto(export_data)
            
        except Exception as e:
            print(f"Error exporting branch deletion report: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def _get_branch_export_data(self, repository_id: int = None) -> List[BranchExportData]:
        """获取分支数据并转换为 DTO"""
        with get_db_session() as db:
            query = db.query(GitlabRepositoryBranch).options(
                joinedload(GitlabRepositoryBranch.repository),
                joinedload(GitlabRepositoryBranch.matched_rule)
            ).join(GitlabRepository)
            
            if repository_id:
                query = query.filter(GitlabRepositoryBranch.repository_id == repository_id)
            
            all_branches = query.all()
            
            # 直接使用数据库中的计算结果，无需重新计算
            export_data = []
            for branch in all_branches:
                # 使用数据库中已计算的结果
                rule_result = {
                    'branch_type': branch.branch_type,
                    'is_deletable': branch.is_deletable,
                    'retention_deadline': branch.retention_deadline,
                    'deletion_reason': branch.deletion_reason,
                    'matched_rule_name': branch.matched_rule.rule_name if branch.matched_rule else None
                }
                dto = BranchExportData.from_model(branch, rule_result)
                export_data.append(dto)
            
            return export_data
    
    def _create_excel_from_dto(self, export_data: List[BranchExportData]) -> io.BytesIO:
        """从 DTO 数据创建 Excel"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # 创建工作表
        self._create_summary_sheet(wb, export_data)
        self._create_deletable_branches_sheet(wb, export_data)
        self._create_all_branches_sheet(wb, export_data)
        
        # 保存到内存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        print(f"Excel file created, size: {len(excel_buffer.getvalue())} bytes")
        return excel_buffer
    
    def _create_summary_sheet(self, wb: Workbook, export_data: List[BranchExportData]):
        """创建汇总页 - 完全使用 DTO"""
        ws = wb.create_sheet("汇总统计")
        
        # 设置标题
        ws['A1'] = '分支删除报告汇总'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # 生成时间
        ws['A3'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # 总体统计 - 使用 DTO 数据
        total_branches = len(export_data)
        deletable_branches = [dto for dto in export_data if dto.computed_deletable]
        protected_branches = [dto for dto in export_data if dto.protected]
        expired_branches = [dto for dto in export_data if dto.is_expired()]
        
        stats = [
            ['统计项', '数量', '占比'],
            ['总分支数', total_branches, '100%'],
            ['可删除分支', len(deletable_branches), f'{len(deletable_branches)/total_branches*100:.1f}%' if total_branches > 0 else '0%'],
            ['受保护分支', len(protected_branches), f'{len(protected_branches)/total_branches*100:.1f}%' if total_branches > 0 else '0%'],
            ['已过期分支', len(expired_branches), f'{len(expired_branches)/total_branches*100:.1f}%' if total_branches > 0 else '0%']
        ]
        
        # 写入统计数据
        for row_idx, row_data in enumerate(stats, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # 标题行
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        # 如果没有分支数据，添加说明
        if total_branches == 0:
            ws['A12'] = '说明：当前没有分支数据，请先同步分支信息'
            ws['A12'].font = Font(color='FF0000')
            return
        
        # 按仓库统计 - 使用 DTO 数据
        repo_stats = {}
        for dto in export_data:
            repo_name = dto.repository_name
            if repo_name not in repo_stats:
                repo_stats[repo_name] = {
                    'total': 0,
                    'deletable': 0,
                    'expired': 0,
                    'protected': 0
                }
            
            repo_stats[repo_name]['total'] += 1
            if dto.computed_deletable:
                repo_stats[repo_name]['deletable'] += 1
            if dto.is_expired():
                repo_stats[repo_name]['expired'] += 1
            if dto.protected:
                repo_stats[repo_name]['protected'] += 1
        
        # 写入仓库统计
        ws['A12'] = '按仓库统计'
        ws['A12'].font = Font(size=14, bold=True)
        
        repo_headers = ['仓库名称', '总分支数', '可删除分支', '已过期分支', '受保护分支']
        for col_idx, header in enumerate(repo_headers, 1):
            cell = ws.cell(row=14, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        for row_idx, (repo_name, stats) in enumerate(repo_stats.items(), 15):
            ws.cell(row=row_idx, column=1, value=repo_name)
            ws.cell(row=row_idx, column=2, value=stats['total'])
            ws.cell(row=row_idx, column=3, value=stats['deletable'])
            ws.cell(row=row_idx, column=4, value=stats['expired'])
            ws.cell(row=row_idx, column=5, value=stats['protected'])
        
        # 按分支类型统计
        type_stats = {}
        for dto in export_data:
            branch_type = dto.computed_type or '未分类'
            if branch_type not in type_stats:
                type_stats[branch_type] = {
                    'total': 0,
                    'deletable': 0,
                    'expired': 0
                }
            
            type_stats[branch_type]['total'] += 1
            if dto.computed_deletable:
                type_stats[branch_type]['deletable'] += 1
            if dto.is_expired():
                type_stats[branch_type]['expired'] += 1
        
        # 写入分支类型统计
        current_row = 15 + len(repo_stats) + 3
        ws.cell(row=current_row, column=1, value='按分支类型统计')
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        
        type_headers = ['分支类型', '总分支数', '可删除分支', '已过期分支']
        for col_idx, header in enumerate(type_headers, 1):
            cell = ws.cell(row=current_row + 2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        for row_idx, (type_name, stats) in enumerate(type_stats.items(), current_row + 3):
            ws.cell(row=row_idx, column=1, value=type_name)
            ws.cell(row=row_idx, column=2, value=stats['total'])
            ws.cell(row=row_idx, column=3, value=stats['deletable'])
            ws.cell(row=row_idx, column=4, value=stats['expired'])
        
        # 自动调整列宽
        self._auto_adjust_column_width(ws)
    
    def _create_deletable_branches_sheet(self, wb: Workbook, export_data: List[BranchExportData]):
        """创建可删除分支明细页 - 完全使用 DTO"""
        ws = wb.create_sheet("可删除分支明细")
        
        # 筛选可删除分支
        deletable_branches = [dto for dto in export_data if dto.computed_deletable]
        
        # 设置标题 - 添加name_with_namespace列
        headers = [
            '仓库名称', '仓库全名', '仓库ID', '分支名称', '分支类型', '最后提交时间', 
            '提交作者', '保留截止时间', '是否已过期', '剩余天数', '删除原因', 
            '是否受保护', '匹配规则', '状态'
        ]
        
        # 写入标题
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if not deletable_branches:
            ws.cell(row=2, column=1, value='暂无可删除的分支')
            ws.cell(row=2, column=1).font = Font(color='FF0000')
            return
        
        # 按过期状态排序，过期的在前面
        deletable_branches.sort(key=lambda x: (not x.is_expired(), x.repository_name, x.branch_name))
        
        # 写入数据 - 使用 DTO 数据，添加name_with_namespace列
        for row_idx, dto in enumerate(deletable_branches, 2):
            ws.cell(row=row_idx, column=1, value=dto.repository_name)
            ws.cell(row=row_idx, column=2, value=dto.name_with_namespace)  # 添加这行
            ws.cell(row=row_idx, column=3, value=dto.repository_id)  # 列索引+1
            ws.cell(row=row_idx, column=4, value=dto.branch_name)    # 列索引+1
            ws.cell(row=row_idx, column=5, value=dto.computed_type or '未分类')  # 列索引+1
            ws.cell(row=row_idx, column=6, value=dto.last_commit_date.strftime('%Y-%m-%d %H:%M:%S') if dto.last_commit_date else '')  # 列索引+1
            ws.cell(row=row_idx, column=7, value=dto.commit_author_name or '')  # 列索引+1
            ws.cell(row=row_idx, column=8, value=dto.computed_deadline.strftime('%Y-%m-%d %H:%M:%S') if dto.computed_deadline else '')  # 列索引+1
            
            # 是否已过期
            is_expired = dto.is_expired()
            expired_cell = ws.cell(row=row_idx, column=9, value='是' if is_expired else '否')  # 列索引+1
            if is_expired:
                expired_cell.font = Font(color='FF0000', bold=True)
            
            # 剩余天数
            days_left = dto.get_days_until_deadline()
            days_cell = ws.cell(row=row_idx, column=10, value=days_left if days_left is not None else '')  # 列索引+1
            if days_left is not None and days_left <= 7:
                days_cell.font = Font(color='FF6600', bold=True)
            
            ws.cell(row=row_idx, column=11, value=dto.computed_reason or '')  # 列索引+1
            ws.cell(row=row_idx, column=12, value='是' if dto.protected else '否')  # 列索引+1
            ws.cell(row=row_idx, column=13, value=dto.matched_rule_name or '')  # 列索引+1
            ws.cell(row=row_idx, column=14, value=dto.get_status_description())  # 列索引+1
        
        # 自动调整列宽
        self._auto_adjust_column_width(ws)
    
    def _create_all_branches_sheet(self, wb: Workbook, export_data: List[BranchExportData]):
        """创建所有分支明细页 - 完全使用 DTO"""
        ws = wb.create_sheet("所有分支明细")
        
        # 设置标题 - 添加name_with_namespace列
        headers = [
            '仓库名称', '仓库全名', '仓库ID', '分支名称', '分支类型', '最后提交时间', 
            '提交作者', '是否受保护', '是否可删除', '保留截止时间', 
            '删除原因', '匹配规则', '状态', '数据来源'
        ]
        
        # 写入标题
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 按仓库和分支名排序
        export_data.sort(key=lambda x: (x.repository_name, x.branch_name))
        
        # 写入数据 - 使用 DTO 数据，添加name_with_namespace列
        for row_idx, dto in enumerate(export_data, 2):
            ws.cell(row=row_idx, column=1, value=dto.repository_name)
            ws.cell(row=row_idx, column=2, value=dto.name_with_namespace)  # 添加这行
            ws.cell(row=row_idx, column=3, value=dto.repository_id)  # 列索引+1
            ws.cell(row=row_idx, column=4, value=dto.branch_name)    # 列索引+1
            ws.cell(row=row_idx, column=5, value=dto.computed_type or '未分类')  # 列索引+1
            ws.cell(row=row_idx, column=6, value=dto.last_commit_date.strftime('%Y-%m-%d %H:%M:%S') if dto.last_commit_date else '')  # 列索引+1
            ws.cell(row=row_idx, column=7, value=dto.commit_author_name or '')  # 列索引+1
            
            # 受保护状态
            protected_cell = ws.cell(row=row_idx, column=8, value='是' if dto.protected else '否')  # 列索引+1
            if dto.protected:
                protected_cell.font = Font(color='0066CC', bold=True)
            
            # 可删除状态
            deletable_cell = ws.cell(row=row_idx, column=9, value='是' if dto.computed_deletable else '否')  # 列索引+1
            if dto.computed_deletable:
                deletable_cell.font = Font(color='FF6600', bold=True)
            
            ws.cell(row=row_idx, column=10, value=dto.computed_deadline.strftime('%Y-%m-%d %H:%M:%S') if dto.computed_deadline else '')  # 列索引+1
            ws.cell(row=row_idx, column=11, value=dto.computed_reason or '')  # 列索引+1
            ws.cell(row=row_idx, column=12, value=dto.matched_rule_name or '')  # 列索引+1
            ws.cell(row=row_idx, column=13, value=dto.get_status_description())  # 列索引+1
            
            # 数据来源标识
            data_source = "计算" if dto.matched_rule_name else "数据库"
            ws.cell(row=row_idx, column=14, value=data_source)  # 列索引+1
        
        # 自动调整列宽
        self._auto_adjust_column_width(ws)
    
    def _auto_adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width